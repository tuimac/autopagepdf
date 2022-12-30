import openpyxl
import json
import sys
import platform
import urllib.request
import logging.config
from time import sleep
from zipfile import ZipFile
from traceback import print_exc, format_exc
from os import chmod, path, _exit, getcwd
from logging import getLogger
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

DRIVER_MAP = {
    "Linux": "chromedriver_linux64.zip",
    "Darwin": "chromedriver_mac64.zip",
    "Windows": "chromedriver_win32.zip"
}
CONFIG_PATH = 'config.json'

def load_conf_file():
    global logger
    global CONF
    try:
        # Loading configration file
        if path.exists('config.json') is False:
            raise FileNotFoundError
        with open('config.json', 'r', encoding='utf-8_sig') as config:
            CONF = json.load(config)

        # Define logger
        logging.config.dictConfig(
            {
                "version": 1,
                "disable_existing_loggers": "True",
                "formatters": {
                    "file": {
                        "format": "[%(levelname)s] - %(asctime)s - Source log file: %(pathname)s line: %(lineno)s %(message)s"
                    }
                },
                "handlers": {
                    "file": {
                        "level": "INFO",
                        "class": "logging.handlers.RotatingFileHandler",
                        "filename": CONF['LOG_FILE_PATH'],
                        "maxBytes": 10485760,
                        "formatter": "file"
                    }
                },
                "loggers": {
                    "main": {
                        "handlers": ["file"],
                        "level": "INFO",
                        "propagate": "True"
                    }
                }
            }
        )
        logger = getLogger('main')
    except:
        print_exc()
        with open(CONF['LOG_FILE_PATH'], 'a', encoding='utf-8_sig') as f:
            f.write(format_exc())
        _exit(1)

def import_excel() -> dict:
    data = dict()
    workbook = openpyxl.load_workbook(CONF['EXCEL_FILE_PATH'], data_only=True)
    sheet = workbook[CONF['EXCEL_SHEET_NAME']]
    row = CONF['EXCEL_DATA_CONFIG']['START_ROW']
    
    while True:
        id_num = sheet.cell(row=row, column=CONF['EXCEL_DATA_CONFIG']['ID_COLUMN']).value
        if id_num == None:
            break
        else:
            url = sheet.cell(row=row, column=CONF['EXCEL_DATA_CONFIG']['URL_COLUMN']).value
            data[id_num] = url
            row += 1
    return data

def __check_exclude_word(url):
    if CONF['EXCLUDE_WORD'] == '':
        return True
    else:
        with urllib.request.urlopen(url) as response:
            result = response.read().decode()
            print(result)
            if CONF['EXCLUDE_WORD'] in result:
                return False
            else:
                return True

def create_pdf(data):
    # Setup Chrome options for prinr page as PDF
    chrome_option = webdriver.ChromeOptions()
    printer_config = {
        'recentDestinations': [
            {
                'id': 'Save as PDF',
                'origin': 'local',
                'account': ''
            }
        ],
        'selectedDestinationId': 'Save as PDF',
        'version': 2,
        'isLandscapeEnabled': False,
        'pageSize': 'A4',
        'marginsType': 0,
        'scalingType': 0,
        'scaling': '100',
        'isHeaderFooterEnabled': False,
        'isCssBackgroundEnabled': True,
        'isDuplexEnabled': False,
        'isColorEnabled': True,
        'isCollateEnabled': True
    }
    
    prefs = {
        'printing.print_preview_sticky_settings.appState': json.dumps(printer_config),
        'download.default_directory': getcwd(),
        'download.open_pdf_in_system_reader': False,
        'download.prompt_for_download': False,
        'plugins.always_open_pdf_externally': True,
        'profile.default_content_settings.popups': 0,

    }

    chrome_option.add_experimental_option('excludeSwitches', ['enable-logging'])
    chrome_option.add_experimental_option('prefs', prefs)
    chrome_option.add_argument('--kiosk-printing')

    # Create PDF from each url website
    driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=chrome_option)

    for key in data:
        try:
            url = data[key]
            if __check_exclude_word(url) is False:
                continue
            driver.get(url)
            WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
            driver.execute_script('document.title="' + str(key) + '";window.print();')
            sleep(CONF['INTERVAL'])
        except urllib.error.HTTPError:
            print_exc()
            logger.error(url + ' is invalid url. So cannot reachable.')
        except:
            print_exc()
            logger.error(format_exc())
    driver.quit()

if __name__ == '__main__':
    try:
        load_conf_file()
        data = import_excel()
        create_pdf(data)
    except:
        print_exc()
        logger.error(format_exc())
